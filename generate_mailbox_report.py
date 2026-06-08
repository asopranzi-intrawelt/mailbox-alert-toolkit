"""
Generatore report mailbox.

Input:
    sys.argv[1] = JSON con dati raccolti oggi
    sys.argv[2] = path output Report_UserMailbox_<data>.xlsx
    sys.argv[3] = path output Report_AltreMailbox_<data>.xlsx
    sys.argv[4] = path SQLite history db (creato se non esiste)

Output:
    - Due file Excel con tutte le colonne richieste
    - Aggiornamento dello storico SQLite (idempotente per data+email)
    - Stampa riepilogo su stdout
"""

import sys
import json
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles  import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils   import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# SCHEMA DB
# ---------------------------------------------------------------------------
SCHEMA = """
CREATE TABLE IF NOT EXISTS mailbox_history (
    snapshot_date     TEXT NOT NULL,
    email             TEXT NOT NULL,
    display_name      TEXT,
    tipo              TEXT,
    licenza           TEXT,
    creata_il         TEXT,
    used_gb           REAL,
    quota_gb          REAL,
    pct_mailbox       REAL,
    numero_email      INTEGER,
    eliminati_gb      REAL,
    inbox_gb          REAL,
    inviati_gb        REAL,
    archivio_abilitato TEXT,
    archivio_used_gb  REAL,
    archivio_quota_gb REAL,
    pct_archivio      REAL,
    ultimo_accesso    TEXT,
    giorni_inattivita INTEGER,
    inoltro_attivo    TEXT,
    litigation_hold   TEXT,
    nascosta_gal      TEXT,
    PRIMARY KEY (snapshot_date, email)
);
CREATE INDEX IF NOT EXISTS idx_email ON mailbox_history(email);
CREATE INDEX IF NOT EXISTS idx_date  ON mailbox_history(snapshot_date);
"""

# ---------------------------------------------------------------------------
# COLONNE EXCEL (ordine, header, larghezza)
# ---------------------------------------------------------------------------
COLUMNS = [
    ("Casella",            "Casella",              38),
    ("DisplayName",        "Display Name",         26),
    ("Tipo",               "Tipo",                 16),
    ("Licenza",            "Licenza",              30),
    ("CreataIl",           "Creata il",            12),
    ("UsatoGB",            "Usato (GB)",           12),
    ("QuotaGB",            "Quota (GB)",           12),
    ("PercMailbox",        "% Occupato",           12),
    ("Crescita30gg",       "Δ 30gg (GB)",          13),
    ("NumeroEmail",        "N° Email",             10),
    ("InboxGB",            "Inbox (GB)",           11),
    ("InviatiGB",          "Inviati (GB)",         12),
    ("EliminatiGB",        "Eliminati (GB)",       13),
    ("ArchivioAbilitato",  "Arch. Att.",           10),
    ("ArchivioUsatoGB",    "Arch. Usato (GB)",     16),
    ("ArchivioQuotaGB",    "Arch. Quota (GB)",     16),
    ("PercArchivio",       "% Archivio",           12),
    ("UltimoAccesso",      "Ultimo Accesso",       18),
    ("GiorniInattivita",   "Inattività (gg)",      15),
    ("InoltroAttivo",      "Inoltro",              12),
    ("LitigationHold",     "Lit. Hold",            10),
    ("NascostaGAL",        "Hidden GAL",           11),
]


def is_archive_enabled(r):
    """Tollerante: matcha 'Sì', 'Si'', 'Si', 'yes', 'true' (case-insensitive)."""
    val = (r.get("ArchivioAbilitato") or "").strip().lower()
    return val.startswith("s") or val in ("yes", "true", "1")

# ---------------------------------------------------------------------------
# HELPERS DI FORMATTAZIONE
# ---------------------------------------------------------------------------
def fill(color):  return PatternFill("solid", start_color=color, fgColor=color)
def thin():       s = Side(style="thin", color="BFBFBF"); return Border(s, s, s, s)

def pct_color(pct):
    if pct is None: return None
    if pct >= 95:   return fill("FFCCCC")
    if pct >= 80:   return fill("FFF3CD")
    return fill("D4EDDA")

def growth_color(delta):
    if delta is None: return None
    if delta >= 2.0:  return fill("FFD6CC")   # crescita rapida
    if delta >= 0.5:  return fill("FFF3CD")
    if delta <= -0.5: return fill("CCE5FF")   # decrescita (cleanup)
    return None

# ---------------------------------------------------------------------------
# SQLITE OPS
# ---------------------------------------------------------------------------
def init_db(db_path):
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.executescript(SCHEMA)
    conn.commit()
    return conn

def upsert_today(conn, rows):
    sql = """
    INSERT OR REPLACE INTO mailbox_history VALUES (
        :snapshot_date,:email,:display_name,:tipo,:licenza,:creata_il,
        :used_gb,:quota_gb,:pct_mailbox,:numero_email,:eliminati_gb,
        :inbox_gb,:inviati_gb,:archivio_abilitato,:archivio_used_gb,
        :archivio_quota_gb,:pct_archivio,:ultimo_accesso,:giorni_inattivita,
        :inoltro_attivo,:litigation_hold,:nascosta_gal
    )
    """
    payload = []
    for r in rows:
        payload.append({
            "snapshot_date":     r.get("Date"),
            "email":             r.get("Casella"),
            "display_name":      r.get("DisplayName"),
            "tipo":              r.get("Tipo"),
            "licenza":           r.get("Licenza"),
            "creata_il":         r.get("CreataIl"),
            "used_gb":           r.get("UsatoGB"),
            "quota_gb":          r.get("QuotaGB"),
            "pct_mailbox":       r.get("PercMailbox"),
            "numero_email":      r.get("NumeroEmail"),
            "eliminati_gb":      r.get("EliminatiGB"),
            "inbox_gb":          r.get("InboxGB"),
            "inviati_gb":        r.get("InviatiGB"),
            "archivio_abilitato":r.get("ArchivioAbilitato"),
            "archivio_used_gb":  r.get("ArchivioUsatoGB"),
            "archivio_quota_gb": r.get("ArchivioQuotaGB"),
            "pct_archivio":      r.get("PercArchivio"),
            "ultimo_accesso":    r.get("UltimoAccesso"),
            "giorni_inattivita": r.get("GiorniInattivita"),
            "inoltro_attivo":    r.get("InoltroAttivo"),
            "litigation_hold":   r.get("LitigationHold"),
            "nascosta_gal":      r.get("NascostaGAL"),
        })
    conn.executemany(sql, payload)
    conn.commit()

def get_growth_30gg(conn, today_iso):
    """Ritorna {email: used_gb_30gg_fa} per calcolare Δ rispetto al passato."""
    target = (datetime.fromisoformat(today_iso) - timedelta(days=30)).date().isoformat()
    # Prendi il record più vicino a 30gg fa (entro ±5gg) per ogni email
    sql = """
    SELECT email, used_gb FROM mailbox_history
    WHERE snapshot_date = (
        SELECT MAX(snapshot_date) FROM mailbox_history h2
        WHERE h2.email = mailbox_history.email
          AND h2.snapshot_date <= ?
          AND h2.snapshot_date >= ?
    )
    """
    five_back = (datetime.fromisoformat(target) - timedelta(days=10)).date().isoformat()
    cur = conn.execute(sql, (target, five_back))
    return {row[0]: row[1] for row in cur.fetchall() if row[1] is not None}

# ---------------------------------------------------------------------------
# GENERAZIONE EXCEL
# ---------------------------------------------------------------------------
def write_sheet(ws, rows, title, report_date, growth_map):
    n_cols = len(COLUMNS)
    last_col_letter = get_column_letter(n_cols)

    # --- Titolo ---
    ws.merge_cells(f"A1:{last_col_letter}1")
    t = ws["A1"]
    t.value = f"{title}  –  {report_date}"
    t.font  = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill  = fill("1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # --- Header ---
    header_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill  = fill("2E75B6")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (_, label, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=2, column=i, value=label)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, thin()
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[2].height = 38

    # --- Dati ---
    base_font = Font(name="Arial", size=9)
    center    = Alignment(horizontal="center", vertical="center")

    for ridx, r in enumerate(rows, start=3):
        # Calcolo crescita 30gg
        prev = growth_map.get(r.get("Casella"))
        if prev is not None and r.get("UsatoGB") is not None:
            r["Crescita30gg"] = round(r["UsatoGB"] - prev, 2)
        else:
            r["Crescita30gg"] = None

        pct = r.get("PercMailbox") or 0
        row_fill = pct_color(pct)

        for cidx, (key, _, _) in enumerate(COLUMNS, start=1):
            val = r.get(key)

            # Conversioni speciali
            if key == "PercMailbox":
                val = (val / 100) if val is not None else None
            elif key == "PercArchivio":
                if not is_archive_enabled(r) or val is None:
                    val = "—"
                else:
                    val = val / 100
            elif key in ("ArchivioUsatoGB","ArchivioQuotaGB"):
                if not is_archive_enabled(r):
                    val = "—"

            cell = ws.cell(row=ridx, column=cidx, value=val)
            cell.font, cell.border, cell.alignment = base_font, thin(), center
            if row_fill: cell.fill = row_fill

            # Override colore archivio
            if key == "PercArchivio" and isinstance(val, float):
                cell.number_format = "0.0%"
                a_fill = pct_color(val * 100)
                if a_fill: cell.fill = a_fill
            elif key == "PercMailbox" and isinstance(val, float):
                cell.number_format = "0.0%"
            elif key == "Crescita30gg" and isinstance(val, (int, float)):
                cell.number_format = '+0.00" GB";-0.00" GB";0" GB"'
                g_fill = growth_color(val)
                if g_fill: cell.fill = g_fill

        ws.row_dimensions[ridx].height = 17

    # --- Riga riepilogo ---
    summary_row = len(rows) + 3
    ws.merge_cells(f"A{summary_row}:{last_col_letter}{summary_row}")
    n_mb_warn = sum(1 for r in rows if (r.get("PercMailbox")  or 0) >= 80)
    n_mb_crit = sum(1 for r in rows if (r.get("PercMailbox")  or 0) >= 95)
    n_ar_warn = sum(1 for r in rows if is_archive_enabled(r) and (r.get("PercArchivio") or 0) >= 80)
    n_ar_crit = sum(1 for r in rows if is_archive_enabled(r) and (r.get("PercArchivio") or 0) >= 95)
    n_arc     = sum(1 for r in rows if is_archive_enabled(r))
    n_inactive = sum(1 for r in rows if (r.get("GiorniInattivita") or 0) >= 90)

    sc = ws.cell(row=summary_row, column=1)
    sc.value = (f"Totale: {len(rows)}   |   "
                f"Mailbox: {n_mb_warn} warning, {n_mb_crit} critical   |   "
                f"Archivio: {n_ar_warn} warning, {n_ar_crit} critical (su {n_arc} con archivio)   |   "
                f"Inattive ≥90gg: {n_inactive}")
    sc.font  = Font(name="Arial", bold=True, size=9, color="1F3864")
    sc.fill  = fill("D9E1F2")
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[summary_row].height = 22

    # Freeze + autofilter
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last_col_letter}{len(rows)+2}"

def build_workbook(path, rows, title, report_date, growth_map):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title.replace(" ", "_")[:31]
    ws.sheet_view.showGridLines = False
    write_sheet(ws, rows, title, report_date, growth_map)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    data_json   = sys.argv[1]
    out_user    = sys.argv[2]
    out_other   = sys.argv[3]
    db_path     = sys.argv[4]

    with open(data_json, "r", encoding="utf-8-sig") as f:
        all_rows = json.load(f)

    # Normalizza in lista anche se è un singolo dict
    if isinstance(all_rows, dict):
        all_rows = [all_rows]

    # SQLite: scrivi storico
    conn = init_db(db_path)
    upsert_today(conn, all_rows)

    today = all_rows[0]["Date"] if all_rows else datetime.now().date().isoformat()
    growth_map = get_growth_30gg(conn, today)

    user_rows  = [r for r in all_rows if r.get("Tipo") == "UserMailbox"]
    other_rows = [r for r in all_rows if r.get("Tipo") != "UserMailbox"]

    rdate = datetime.now().strftime("%d/%m/%Y %H:%M")

    build_workbook(out_user,  user_rows,  "Report User Mailbox",  rdate, growth_map)
    build_workbook(out_other, other_rows, "Report Altre Mailbox", rdate, growth_map)

    conn.close()
    print(f"OK | User: {len(user_rows)} | Altre: {len(other_rows)} | "
          f"Storico righe: {sum(1 for _ in all_rows)} | Growth confronti: {len(growth_map)}")

if __name__ == "__main__":
    main()
