"""
Generatore TRENDS - analytics storiche dalla history SQLite.

Input:
    sys.argv[1] = path SQLite history db
    sys.argv[2] = path output Trends.xlsx

Produce un workbook con i seguenti fogli:
    1. "Riepilogo Tenant"            - serie temporale tenant (mailbox + archivio)
    2. "Top Crescita Mailbox 30gg"   - top 30 caselle per crescita mailbox negli ultimi 30 giorni
    3. "Top Crescita Mailbox 12mesi" - top 30 caselle per crescita mailbox negli ultimi 12 mesi
    4. "Top Crescita Archivio 30gg"  - top 30 caselle per crescita archivio negli ultimi 30 giorni
    5. "Top Crescita Archivio 12mesi"- top 30 caselle per crescita archivio negli ultimi 12 mesi
    6. "Inattive"                     - caselle con >= 90gg di inattivita' e spazio occupato
    7. "Storico Mailbox per Casella"  - pivot data x email -> GB mailbox
    8. "Storico Archivio per Casella" - pivot data x email -> GB archivio

Note matematiche:
- Il riferimento "oggi" e' MAX(snapshot_date) nel DB, non datetime.now(),
  per evitare disallineamenti quando il task non gira da qualche giorno.
- "Delta medio per giorno" e' calcolato usando i giorni EFFETTIVI tra il
  primo e l'ultimo snapshot trovato per quella casella nella finestra, non
  un valore nominale (30 o 365).
"""

import sys
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles  import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils   import get_column_letter
from openpyxl.chart   import LineChart, Reference

# ---------------------------------------------------------------------------
# STYLE HELPERS
# ---------------------------------------------------------------------------
def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def thin():  s = Side(style="thin", color="BFBFBF"); return Border(s, s, s, s)

HDR_FILL_MAIL = fill("2E75B6")   # blu per mailbox
HDR_FILL_ARC  = fill("7B4F9F")   # viola per archivio
TITLE_FILL    = fill("1F3864")
HDR_FONT      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
BASE_FONT     = Font(name="Arial", size=9)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)

def write_title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font, c.fill, c.alignment = Font(name="Arial", bold=True, size=13, color="FFFFFF"), TITLE_FILL, CENTER
    ws.row_dimensions[1].height = 30

def write_headers(ws, headers, widths, row=2, hdr_fill=None):
    if hdr_fill is None: hdr_fill = HDR_FILL_MAIL
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = HDR_FONT, hdr_fill, CENTER, thin()
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1] if i-1 < len(widths) else 15
    ws.row_dimensions[row].height = 32

def write_data_cell(ws, row, col, value, number_format=None, fill_color=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font, c.alignment, c.border = BASE_FONT, CENTER, thin()
    if number_format: c.number_format = number_format
    if fill_color:    c.fill = fill_color
    return c

# ---------------------------------------------------------------------------
# FOGLIO 1: Riepilogo tenant (serie temporale, mailbox + archivio)
# ---------------------------------------------------------------------------
def sheet_tenant_summary(wb, conn):
    ws = wb.create_sheet("Riepilogo Tenant")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Riepilogo Tenant - Serie Temporale (mailbox e archivio)", 11)
    write_headers(ws,
        ["Data",
         "N. Caselle",
         "Tot. Mailbox (GB)", "% Media Mb", "Mb >=80%", "Mb >=95%",
         "N. Archivi", "Tot. Archivio (GB)", "% Media Arc", "Arc >=80%", "Arc >=95%"],
        [12, 11, 17, 13, 11, 11, 12, 17, 13, 11, 11])

    sql = """
    SELECT snapshot_date,
           COUNT(*)                                        AS n,
           ROUND(SUM(used_gb), 2)                          AS tot_mb,
           ROUND(AVG(pct_mailbox), 1)                      AS avg_mb_pct,
           SUM(CASE WHEN pct_mailbox>=80 THEN 1 ELSE 0 END) AS n_mb_80,
           SUM(CASE WHEN pct_mailbox>=95 THEN 1 ELSE 0 END) AS n_mb_95,
           SUM(CASE WHEN LOWER(COALESCE(archivio_abilitato,'')) LIKE 's%' THEN 1 ELSE 0 END) AS n_arc,
           ROUND(SUM(CASE WHEN LOWER(COALESCE(archivio_abilitato,'')) LIKE 's%'
                          THEN archivio_used_gb ELSE 0 END), 2) AS tot_arc,
           ROUND(AVG(CASE WHEN LOWER(COALESCE(archivio_abilitato,'')) LIKE 's%'
                          THEN pct_archivio END), 1) AS avg_arc_pct,
           SUM(CASE WHEN pct_archivio>=80 THEN 1 ELSE 0 END) AS n_arc_80,
           SUM(CASE WHEN pct_archivio>=95 THEN 1 ELSE 0 END) AS n_arc_95
    FROM mailbox_history
    GROUP BY snapshot_date
    ORDER BY snapshot_date
    """
    rows = conn.execute(sql).fetchall()
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            write_data_cell(ws, r_idx, c_idx, val)
    ws.freeze_panes = "A3"

    # Grafico spazio totale mailbox vs archivio nel tempo (se dataset > 1 giorno)
    if len(rows) > 1:
        chart = LineChart()
        chart.title = "Spazio Totale Tenant nel Tempo (GB)"
        chart.y_axis.title = "GB"
        chart.x_axis.title = "Data"
        # Serie mailbox (col 3) e archivio (col 8)
        ref_mb  = Reference(ws, min_col=3, min_row=2, max_row=len(rows)+2)
        ref_arc = Reference(ws, min_col=8, min_row=2, max_row=len(rows)+2)
        cats    = Reference(ws, min_col=1, min_row=3, max_row=len(rows)+2)
        chart.add_data(ref_mb,  titles_from_data=True)
        chart.add_data(ref_arc, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 11; chart.width = 24
        ws.add_chart(chart, "M3")

# ---------------------------------------------------------------------------
# TOP CRESCITA (generico, parametrico su mailbox o archivio)
# ---------------------------------------------------------------------------
def sheet_top_growth(wb, conn, last_snap, days, sheet_name, kind):
    """
    kind: 'mailbox' o 'archivio'
    """
    if kind == "mailbox":
        usage_col   = "used_gb"
        used_label  = "Mailbox"
        hdr_fill    = HDR_FILL_MAIL
        # nessun filtro extra
        extra_where = ""
    elif kind == "archivio":
        usage_col   = "archivio_used_gb"
        used_label  = "Archivio"
        hdr_fill    = HDR_FILL_ARC
        # filtra solo caselle con archivio attivo
        extra_where = "AND LOWER(COALESCE(m1.archivio_abilitato,'')) LIKE 's%'"
    else:
        raise ValueError(f"kind sconosciuto: {kind}")

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    write_title(ws, f"Top 30 caselle - Crescita {used_label.upper()} ultimi {days} giorni", 9)
    write_headers(ws,
        ["Casella", "Tipo",
         "Data Inizio", f"Inizio ({used_label}, GB)",
         "Data Fine",   f"Fine ({used_label}, GB)",
         "Δ (GB)", "Δ %", "Δ medio/gg (MB)"],
        [38, 16, 12, 17, 12, 17, 11, 9, 16],
        hdr_fill=hdr_fill)

    # Target a `days` giorni indietro rispetto all'ULTIMO snapshot, non a now()
    target = (datetime.fromisoformat(last_snap) - timedelta(days=days)).date().isoformat()

    sql = f"""
    WITH latest AS (
        SELECT email, MAX(snapshot_date) AS d
        FROM mailbox_history
        WHERE {usage_col} IS NOT NULL
        GROUP BY email
    ),
    oldest AS (
        SELECT email, MIN(snapshot_date) AS d
        FROM mailbox_history
        WHERE snapshot_date >= ?
          AND {usage_col} IS NOT NULL
        GROUP BY email
    )
    SELECT  m1.email, m1.tipo,
            m_old.snapshot_date                  AS data_inizio,
            ROUND(m_old.{usage_col}, 2)          AS gb_start,
            m1.snapshot_date                     AS data_fine,
            ROUND(m1.{usage_col}, 2)             AS gb_end,
            ROUND(m1.{usage_col} - m_old.{usage_col}, 2) AS delta_gb,
            CASE WHEN m_old.{usage_col} > 0
                 THEN ROUND((m1.{usage_col} - m_old.{usage_col}) * 100.0 / m_old.{usage_col}, 1)
                 ELSE NULL END                   AS delta_pct
    FROM   mailbox_history m1
    JOIN   latest l ON l.email = m1.email AND l.d = m1.snapshot_date
    JOIN   oldest o ON o.email = m1.email
    JOIN   mailbox_history m_old
           ON m_old.email = o.email
           AND m_old.snapshot_date = o.d
    WHERE  m_old.{usage_col} IS NOT NULL
      AND  m1.{usage_col} IS NOT NULL
      {extra_where}
    ORDER BY delta_gb DESC
    LIMIT 30
    """
    rows = conn.execute(sql, (target,)).fetchall()

    if not rows:
        ws.cell(row=3, column=1,
                value=f"Nessun dato disponibile per crescita {used_label.lower()} negli ultimi {days} giorni.")
        return

    for r_idx, row in enumerate(rows, start=3):
        email, tipo, d_start, gb_start, d_end, gb_end, delta_gb, delta_pct = row

        # giorni effettivi tra inizio e fine (non il valore nominale `days`)
        try:
            actual_days = (datetime.fromisoformat(d_end) - datetime.fromisoformat(d_start)).days
        except Exception:
            actual_days = 0
        actual_days = max(1, actual_days)
        mb_day = round((delta_gb or 0) * 1024 / actual_days, 1)

        # evidenzia crescite forti
        row_fill = None
        if delta_gb is not None and delta_gb >= 5:
            row_fill = fill("FFE5CC")
        elif delta_gb is not None and delta_gb <= -1:
            row_fill = fill("D6EAF8")

        for c_idx, val in enumerate(
                [email, tipo, d_start, gb_start, d_end, gb_end, delta_gb, delta_pct, mb_day],
                start=1):
            write_data_cell(ws, r_idx, c_idx, val, fill_color=row_fill)

    ws.freeze_panes = "A3"

# ---------------------------------------------------------------------------
# FOGLIO INATTIVE
# ---------------------------------------------------------------------------
def sheet_inactive(wb, conn):
    ws = wb.create_sheet("Inattive")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Caselle inattive (>= 90 giorni) - Mailbox e Archivio", 8)
    write_headers(ws,
        ["Casella", "Tipo", "Licenza",
         "Mailbox (GB)", "Archivio (GB)",
         "Giorni Inattivita'", "Ultimo Accesso", "Snapshot"],
        [38, 16, 28, 12, 13, 16, 18, 12])

    sql = """
    SELECT email, tipo, licenza,
           used_gb,
           CASE WHEN LOWER(COALESCE(archivio_abilitato,'')) LIKE 's%'
                THEN archivio_used_gb ELSE NULL END AS arc_used,
           giorni_inattivita, ultimo_accesso, snapshot_date
    FROM mailbox_history
    WHERE snapshot_date = (SELECT MAX(snapshot_date) FROM mailbox_history)
      AND giorni_inattivita >= 90
    ORDER BY (used_gb + COALESCE(archivio_used_gb,0)) DESC
    """
    rows = conn.execute(sql).fetchall()
    if not rows:
        ws.cell(row=3, column=1, value="Nessuna casella con piu' di 90 giorni di inattivita'.")
        return

    for r_idx, row in enumerate(rows, start=3):
        used_gb  = row[3] or 0
        arc_used = row[4] or 0
        total    = used_gb + arc_used
        # evidenzia se inattiva ma occupa > 5GB totali
        row_fill = fill("FFF3CD") if total > 5 else None
        for c_idx, val in enumerate(row, start=1):
            write_data_cell(ws, r_idx, c_idx, val, fill_color=row_fill)
    ws.freeze_panes = "A3"

# ---------------------------------------------------------------------------
# STORICO PIVOT (mailbox e archivio in fogli separati)
# ---------------------------------------------------------------------------
def sheet_pivot(wb, conn, sheet_name, kind):
    """
    kind: 'mailbox' o 'archivio'
    """
    if kind == "mailbox":
        col = "used_gb"
        title = "Storico Mailbox (GB) - una riga per casella, una colonna per snapshot"
        extra = ""
    else:
        col = "archivio_used_gb"
        title = "Storico Archivio (GB) - solo caselle con archivio attivo"
        extra = "AND LOWER(COALESCE(archivio_abilitato,'')) LIKE 's%'"

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    dates = [r[0] for r in conn.execute(
        "SELECT DISTINCT snapshot_date FROM mailbox_history ORDER BY snapshot_date"
    ).fetchall()]
    emails = [r[0] for r in conn.execute(
        f"SELECT DISTINCT email FROM mailbox_history WHERE 1=1 {extra} ORDER BY email"
    ).fetchall()]

    write_title(ws, f"{title}  -  {len(emails)} caselle x {len(dates)} snapshot",
                len(dates) + 1)

    # Header riga 2
    c = ws.cell(row=2, column=1, value="Casella")
    c.font, c.fill, c.alignment, c.border = HDR_FONT, (HDR_FILL_MAIL if kind=="mailbox" else HDR_FILL_ARC), CENTER, thin()
    ws.column_dimensions["A"].width = 38
    for i, d in enumerate(dates, start=2):
        c = ws.cell(row=2, column=i, value=d)
        c.font, c.fill, c.alignment, c.border = HDR_FONT, (HDR_FILL_MAIL if kind=="mailbox" else HDR_FILL_ARC), CENTER, thin()
        ws.column_dimensions[get_column_letter(i)].width = 11

    if not emails or not dates:
        ws.cell(row=3, column=1, value="Nessun dato disponibile.")
        return

    # Carico tutti i dati in una query
    sql = f"SELECT email, snapshot_date, {col} FROM mailbox_history WHERE 1=1 {extra}"
    data = {}
    for email, d, v in conn.execute(sql).fetchall():
        data.setdefault(email, {})[d] = v

    for r_idx, email in enumerate(emails, start=3):
        c = ws.cell(row=r_idx, column=1, value=email)
        c.font, c.alignment, c.border = BASE_FONT, Alignment(horizontal="left", indent=1), thin()
        for i, d in enumerate(dates, start=2):
            v = data.get(email, {}).get(d)
            cell = ws.cell(row=r_idx, column=i, value=v)
            cell.font, cell.alignment, cell.border = BASE_FONT, CENTER, thin()
            if isinstance(v, (int, float)) and v > 0:
                cell.number_format = "0.00"

    ws.freeze_panes = "B3"

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    if len(sys.argv) < 3:
        print("Uso: generate_trends.py <history.db> <output.xlsx>")
        sys.exit(2)

    db_path  = sys.argv[1]
    out_path = sys.argv[2]

    if not Path(db_path).exists():
        print(f"ERRORE: nessuno storico trovato in {db_path}")
        sys.exit(1)

    conn = sqlite3.connect(db_path)

    last_snap = conn.execute("SELECT MAX(snapshot_date) FROM mailbox_history").fetchone()[0]
    if not last_snap:
        print("ERRORE: il database storico e' vuoto.")
        sys.exit(1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # rimuovi foglio default

    sheet_tenant_summary(wb, conn)
    sheet_top_growth(wb, conn, last_snap, days=30,  sheet_name="Top Crescita Mailbox 30gg",   kind="mailbox")
    sheet_top_growth(wb, conn, last_snap, days=365, sheet_name="Top Crescita Mailbox 12mesi", kind="mailbox")
    sheet_top_growth(wb, conn, last_snap, days=30,  sheet_name="Top Crescita Archivio 30gg",  kind="archivio")
    sheet_top_growth(wb, conn, last_snap, days=365, sheet_name="Top Crescita Archivio 12mesi",kind="archivio")
    sheet_inactive(wb, conn)
    sheet_pivot(wb, conn, "Storico Mailbox per Casella",  kind="mailbox")
    sheet_pivot(wb, conn, "Storico Archivio per Casella", kind="archivio")

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    conn.close()
    print(f"OK | Trends report: {out_path} | Ultimo snapshot riferimento: {last_snap}")

if __name__ == "__main__":
    main()
